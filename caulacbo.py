from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from Database.data import get_db_connection


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        self.MainWindow = MainWindow
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(635, 583)
        MainWindow.setStyleSheet("""
            QPushButton {
                background-color: #4A90E2;
                color: #FFFFFF;
                font-weight: bold;
                border: 1px solid #357ABD;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #357ABD;
                color: #FFFFFF;
                border: 2px solid #4A90E2;
            }
            QLabel#titleLabel {
                font-size: 20px;
                font-weight: bold;
                color: #357ABD;
            }
            QWidget {
                background-color: #E0E0E0;
                color: #333333;
                font-weight: bold;
            }
            QLineEdit {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
                border-radius: 3px;
                padding: 2px;
            }
            QComboBox {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
                border-radius: 3px;
                padding: 2px;
            }
            QGroupBox {
                font-weight: bold;
                color: #357ABD;
            }
            QTableWidget {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
            }
            QTableWidget::item:selected {
                background-color: #4A90E2;
                color: #FFFFFF;
            }
        """)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.lineEdit = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(0, -10, 661, 571))
        self.lineEdit.setStyleSheet(
            "QWidget { background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #E1F5FE, stop:1 #B3E5FC); }")
        self.lineEdit.setObjectName("lineEdit")

        self.groupBox_6 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_6.setGeometry(QtCore.QRect(20, 270, 601, 61))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.groupBox_6.setFont(font)
        self.groupBox_6.setObjectName("groupBox_6")

        self.btnthem = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnthem.setGeometry(QtCore.QRect(50, 20, 93, 28))
        self.btnthem.setObjectName("btnthem")

        self.btnlammoi = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnlammoi.setGeometry(QtCore.QRect(370, 20, 93, 28))
        self.btnlammoi.setObjectName("btnlammoi")

        self.btnluu = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnluu.setGeometry(QtCore.QRect(150, 20, 93, 28))
        self.btnluu.setObjectName("btnluu")

        self.btnxoa = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnxoa.setGeometry(QtCore.QRect(260, 20, 93, 28))
        self.btnxoa.setObjectName("btnxoa")

        self.btnthoat = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnthoat.setGeometry(QtCore.QRect(480, 20, 93, 28))
        self.btnthoat.setObjectName("btnthoat")

        self.groupBox = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(20, 60, 311, 201))
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")

        self.label_5 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(10, 40, 91, 16))
        self.label_5.setObjectName("label_5")

        self.label_11 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_11.setGeometry(QtCore.QRect(10, 80, 81, 16))
        self.label_11.setObjectName("label_11")

        self.label_15 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_15.setGeometry(QtCore.QRect(10, 120, 91, 16))
        self.label_15.setObjectName("label_15")

        self.txtmaclb = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtmaclb.setGeometry(QtCore.QRect(100, 40, 161, 22))
        self.txtmaclb.setObjectName("txtmaclb")

        self.label_16 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_16.setGeometry(QtCore.QRect(10, 160, 91, 16))
        self.label_16.setObjectName("label_16")

        self.txtmota = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtmota.setGeometry(QtCore.QRect(100, 160, 161, 22))
        self.txtmota.setObjectName("txtmota")

        self.txtdiachi = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtdiachi.setGeometry(QtCore.QRect(100, 120, 161, 22))
        self.txtdiachi.setObjectName("txtdiachi")

        self.txttenclb = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txttenclb.setGeometry(QtCore.QRect(100, 80, 161, 22))
        self.txttenclb.setObjectName("txttenclb")

        self.groupBox_2 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_2.setGeometry(QtCore.QRect(20, 340, 601, 211))
        self.groupBox_2.setFont(font)
        self.groupBox_2.setObjectName("groupBox_2")

        self.tblds = QtWidgets.QTableWidget(parent=self.groupBox_2)
        self.tblds.setGeometry(QtCore.QRect(10, 20, 581, 181))
        self.tblds.setColumnCount(4)
        self.tblds.setHorizontalHeaderLabels(["Mã CLB", "Tên CLB", "Địa Chỉ", "Mô Tả"])
        self.tblds.setColumnWidth(0, 60)
        self.tblds.setColumnWidth(1, 150)
        self.tblds.setColumnWidth(2, 170)
        self.tblds.setColumnWidth(3, 250)
        self.tblds.setObjectName("tblds")

        self.groupBox_3 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(340, 60, 281, 201))
        self.groupBox_3.setFont(font)
        self.groupBox_3.setObjectName("groupBox_3")

        self.label_20 = QtWidgets.QLabel(parent=self.groupBox_3)
        self.label_20.setGeometry(QtCore.QRect(50, 30, 101, 16))
        self.label_20.setObjectName("label_20")

        self.txtmaclb_2 = QtWidgets.QComboBox(parent=self.groupBox_3)
        self.txtmaclb_2.setGeometry(QtCore.QRect(50, 50, 151, 21))
        self.txtmaclb_2.setObjectName("txtmaclb_2")

        self.btnxuat = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btnxuat.setGeometry(QtCore.QRect(160, 150, 81, 28))
        self.btnxuat.setObjectName("btnxuat")

        self.btnnhap = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btnnhap.setGeometry(QtCore.QRect(60, 150, 81, 28))
        self.btnnhap.setObjectName("btnnhap")

        self.btntimkiem = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btntimkiem.setGeometry(QtCore.QRect(110, 100, 81, 31))
        self.btntimkiem.setFont(font)
        self.btntimkiem.setObjectName("btntimkiem")

        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        self.label.setGeometry(QtCore.QRect(190, 10, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet(
            "color: #4A90E2; background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #E1F5FE, stop:1 #B3E5FC);")
        self.label.setObjectName("label")

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.btnthem.clicked.connect(self.add_account)
        self.btnxoa.clicked.connect(self.delete_account)
        self.btnluu.clicked.connect(self.update_account)
        self.btnxuat.clicked.connect(self.export_data)
        self.btnnhap.clicked.connect(self.import_data)
        self.btnlammoi.clicked.connect(self.lam_moi_du_lieu)
        self.btntimkiem.clicked.connect(self.search_caulacbo)
        self.btnthoat.clicked.connect(self.thoat)
        self.tblds.itemSelectionChanged.connect(self.table_item_selected)

        self.load_data()
        self.load_account_ids()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.groupBox_6.setTitle(_translate("MainWindow", "Xử lý"))
        self.btnthem.setText(_translate("MainWindow", "Thêm"))
        self.btnlammoi.setText(_translate("MainWindow", "Làm mới"))
        self.btnluu.setText(_translate("MainWindow", "Lưu"))
        self.btnxoa.setText(_translate("MainWindow", "Xóa"))
        self.btnthoat.setText(_translate("MainWindow", "Thoát"))
        self.groupBox.setTitle(_translate("MainWindow", "Thông tin câu lạc bộ:"))
        self.label_5.setText(_translate("MainWindow", "Mã CLB:"))
        self.label_11.setText(_translate("MainWindow", "Tên CLB:"))
        self.label_15.setText(_translate("MainWindow", "Địa chỉ:"))
        self.label_16.setText(_translate("MainWindow", "Mô tả:"))
        self.groupBox_2.setTitle(_translate("MainWindow", "Danh sách câu lạc bộ:"))
        self.groupBox_3.setTitle(_translate("MainWindow", "Tìm kiếm"))
        self.label_20.setText(_translate("MainWindow", "Tên CLB:"))
        self.btnxuat.setText(_translate("MainWindow", "Xuất Excel"))
        self.btnnhap.setText(_translate("MainWindow", "Nhập Excel"))
        self.btntimkiem.setText(_translate("MainWindow", "Tìm kiếm"))
        self.label.setText(_translate("MainWindow", "QUẢN LÝ CÂU LẠC BỘ"))

    def import_data(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(None, "Chọn file Excel", "", "Excel Files (*.xlsx)")
            if not file_path:
                return

            wb = load_workbook(file_path)
            ws = wb.active

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            inserted_count = 0
            duplicate_maclbs = []  # Danh sách lưu các mã CLB bị trùng

            # Bỏ qua dòng tiêu đề (dòng 1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                maclb, tenclb, diachi, mota = row

                # Kiểm tra dữ liệu hợp lệ
                if not maclb or not tenclb:
                    continue

                # Kiểm tra mã CLB đã tồn tại chưa
                cursor.execute("SELECT COUNT(*) FROM caulacbo WHERE ma_clb = %s", (maclb,))
                if cursor.fetchone()[0] > 0:
                    duplicate_maclbs.append(str(maclb))  # Thêm mã bị trùng vào danh sách
                    continue

                # Thêm dữ liệu vào database
                cursor.execute("""
                    INSERT INTO caulacbo (ma_clb, ten_clb, dia_chi, mo_ta)
                    VALUES (%s, %s, %s, %s)
                """, (maclb, tenclb, diachi, mota))
                inserted_count += 1

            connection.commit()
            self.load_data()
            # Tạo thông báo kết quả
            message = f"Đã nhập thành công {inserted_count} câu lạc bộ từ Excel!"
            if duplicate_maclbs:
                message += f"\nCác mã CLB bị trùng (không được nhập): {', '.join(duplicate_maclbs)}"
            QMessageBox.information(None, "Kết quả nhập", message)

            self.load_data()  # Cập nhật bảng
            self.load_account_ids()  # Cập nhật combobox
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi nhập: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def load_data(self):
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM caulacbo")
            rows = cursor.fetchall()
            self.tblds.setRowCount(0)
            for row in rows:
                rowPosition = self.tblds.rowCount()
                self.tblds.insertRow(rowPosition)
                for col, value in enumerate(row):
                    self.tblds.setItem(rowPosition, col, QtWidgets.QTableWidgetItem(str(value)))
            cursor.close()
            connection.close()

    def load_account_ids(self):
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT ten_clb FROM caulacbo")
            rows = cursor.fetchall()
            self.txtmaclb_2.clear()
            for row in rows:
                self.txtmaclb_2.addItem(row[0])
            cursor.close()
            connection.close()

    def search_caulacbo(self):
        tenclb = self.txtmaclb_2.currentText()
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM caulacbo WHERE ten_clb = %s", (tenclb,))
            rows = cursor.fetchall()
            self.tblds.setRowCount(0)
            for row in rows:
                rowPosition = self.tblds.rowCount()
                self.tblds.insertRow(rowPosition)
                for col, value in enumerate(row):
                    self.tblds.setItem(rowPosition, col, QtWidgets.QTableWidgetItem(str(value)))
            cursor.close()
            connection.close()

    def add_account(self):
        try:
            maclb = self.txtmaclb.text().strip()
            tenclb = self.txttenclb.text().strip()
            diachi = self.txtdiachi.text().strip()
            mota = self.txtmota.text().strip()

            if not maclb or not tenclb:
                QMessageBox.warning(None, "Lỗi", "Mã CLB và Tên CLB không được để trống!")
                return

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM caulacbo WHERE ma_clb = %s", (maclb,))
            count = cursor.fetchone()[0]

            if count > 0:
                QMessageBox.warning(None, "Lỗi", "Mã CLB đã tồn tại!")
                cursor.close()
                connection.close()
                return

            cursor.execute("""
                INSERT INTO caulacbo (ma_clb, ten_clb, dia_chi, mo_ta)
                VALUES (%s, %s, %s, %s)
            """, (maclb, tenclb, diachi, mota))
            connection.commit()
            QMessageBox.information(None, "Thành công", "Thêm câu lạc bộ thành công!")
            self.load_data()
            self.load_account_ids()
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi thêm: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def delete_account(self):
        try:
            row = self.tblds.currentRow()
            if row == -1:
                QMessageBox.warning(None, "Lỗi", "Vui lòng chọn một dòng để xóa!")
                return

            maclb = self.tblds.item(row, 0).text()
            reply = QMessageBox.question(None, "Xác nhận", "Bạn có chắc muốn xóa câu lạc bộ này?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                connection = get_db_connection()
                if not connection:
                    QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                    return

                cursor = connection.cursor()
                cursor.execute("DELETE FROM caulacbo WHERE ma_clb = %s", (maclb,))
                connection.commit()
                QMessageBox.information(None, "Thành công", "Xóa câu lạc bộ thành công!")
                self.load_data()
                self.load_account_ids()
                cursor.close()
                connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi xóa: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def update_account(self):
        try:
            row = self.tblds.currentRow()
            if row == -1:
                QMessageBox.warning(None, "Lỗi", "Vui lòng chọn một dòng để sửa!")
                return

            maclb = self.txtmaclb.text().strip()
            tenclb = self.txttenclb.text().strip()
            diachi = self.txtdiachi.text().strip()
            mota = self.txtmota.text().strip()

            if not maclb or not tenclb:
                QMessageBox.warning(None, "Lỗi", "Mã CLB và Tên CLB không được để trống!")
                return

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            cursor.execute("""
                UPDATE caulacbo SET ten_clb = %s, dia_chi = %s, mo_ta = %s
                WHERE ma_clb = %s
            """, (tenclb, diachi, mota, maclb))
            connection.commit()
            QMessageBox.information(None, "Thành công", "Cập nhật câu lạc bộ thành công!")
            self.load_data()
            self.load_account_ids()
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi sửa: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def lam_moi_du_lieu(self):
        self.txtmaclb.clear()
        self.txttenclb.clear()
        self.txtdiachi.clear()
        self.txtmota.clear()
        self.txtmaclb_2.setCurrentIndex(-1)
        self.load_data()
        self.load_account_ids()
        QMessageBox.information(None, "Thành công", "Dữ liệu đã được làm mới!")

    def thoat(self):
        reply = QMessageBox.question(None, "Xác nhận", "Bạn có muốn thoát không?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.MainWindow.close()

    def export_data(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "CauLacBo Data"
        headers = ["Mã CLB", "Tên CLB", "Địa Chỉ", "Mô Tả"]
        ws.append(headers)

        for row in range(self.tblds.rowCount()):
            row_data = []
            for col in range(self.tblds.columnCount()):
                item = self.tblds.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)

        file_path, _ = QFileDialog.getSaveFileName(None, "Lưu file Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QMessageBox.information(None, "Thành công", "Dữ liệu đã được xuất thành công!")

    def table_item_selected(self):
        selected_row = self.tblds.currentRow()
        if selected_row != -1:
            self.txtmaclb.setText(self.tblds.item(selected_row, 0).text())
            self.txttenclb.setText(self.tblds.item(selected_row, 1).text())
            self.txtdiachi.setText(self.tblds.item(selected_row, 2).text())
            self.txtmota.setText(self.tblds.item(selected_row, 3).text())


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())