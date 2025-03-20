from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from Database.data import get_db_connection


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        self.MainWindow = MainWindow
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(822, 588)
        MainWindow.setStyleSheet("""
            QPushButton {
                background-color: #4A90E2;
                color: #FFFFFF;
                font-weight: bold;
                border: 1px solid #357ABD;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #357ABD;
                color: #FFFFFF;
                border: 2px solid #4A90E2;
            }
            QLabel#titleLabel {
                font-size: 20px;
                font-weight: bold;
                color: #357ABD;
            }
            QWidget {
                background-color: #E0E0E0;
                color: #333333;
                font-weight: bold;
            }
            QLineEdit {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
                border-radius: 3px;
                padding: 2px;
            }
            QComboBox, QDateEdit {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
                border-radius: 3px;
                padding: 2px;
            }
            QGroupBox {
                font-weight: bold;
                color: #357ABD;
                font-size: 12px;
            }
            QTableWidget {
                background-color: #FFFFFF;
                color: #333333;
                border: 1px solid #B0B0B0;
            }
            QTableWidget::item:selected {
                background-color: #4A90E2;
                color: #FFFFFF;
            }
        """)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.lineEdit = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(0, 0, 821, 561))
        self.lineEdit.setStyleSheet(
            "QWidget { background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #E1F5FE, stop:1 #B3E5FC); }")

        # GroupBox tìm kiếm
        self.groupBox_3 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(620, 60, 181, 201))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setPointSize(10)
        self.groupBox_3.setFont(font)
        self.label_20 = QtWidgets.QLabel(parent=self.groupBox_3)
        self.label_20.setGeometry(QtCore.QRect(10, 30, 101, 16))
        self.txtmatv_2 = QtWidgets.QComboBox(parent=self.groupBox_3)
        self.txtmatv_2.setGeometry(QtCore.QRect(10, 50, 151, 22))

        # Sắp xếp 3 nút theo hàng dọc
        self.btntimkiem = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btntimkiem.setGeometry(QtCore.QRect(40, 80, 91, 31))
        self.btntimkiem.setFont(font)

        self.btnnhap = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btnnhap.setGeometry(QtCore.QRect(40, 120, 91, 31))

        self.btnxuat = QtWidgets.QPushButton(parent=self.groupBox_3)
        self.btnxuat.setGeometry(QtCore.QRect(40, 160, 91, 31))

        # GroupBox thông tin
        self.groupBox = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(10, 60, 601, 201))
        self.groupBox.setFont(font)
        self.label_5 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(10, 40, 91, 16))
        self.label_11 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_11.setGeometry(QtCore.QRect(10, 80, 55, 16))
        self.label_15 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_15.setGeometry(QtCore.QRect(10, 120, 91, 16))
        self.txtmatv = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtmatv.setGeometry(QtCore.QRect(100, 40, 161, 22))
        self.label_17 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_17.setGeometry(QtCore.QRect(290, 40, 55, 16))
        self.label_18 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_18.setGeometry(QtCore.QRect(290, 120, 81, 16))
        self.label_19 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_19.setGeometry(QtCore.QRect(290, 80, 91, 16))
        self.txtdiachi = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtdiachi.setGeometry(QtCore.QRect(390, 40, 171, 22))
        self.txtsdt = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txtsdt.setGeometry(QtCore.QRect(390, 80, 171, 22))
        self.txtngaysinh = QtWidgets.QDateEdit(parent=self.groupBox)
        self.txtngaysinh.setGeometry(QtCore.QRect(100, 120, 161, 22))
        self.txtngaysinh.setCalendarPopup(True)
        self.txtngaysinh.setDate(QtCore.QDate.currentDate())
        self.txtngaysinh.setDisplayFormat("dd/MM/yyyy")
        self.txtmaclb = QtWidgets.QComboBox(parent=self.groupBox)
        self.txtmaclb.setGeometry(QtCore.QRect(390, 120, 171, 22))
        self.label_16 = QtWidgets.QLabel(parent=self.groupBox)
        self.label_16.setGeometry(QtCore.QRect(10, 160, 91, 16))
        self.txtgioitinh = QtWidgets.QComboBox(parent=self.groupBox)
        self.txtgioitinh.setGeometry(QtCore.QRect(100, 160, 161, 22))
        self.txtgioitinh.addItems(["Nam", "Nữ"])
        self.txttentv = QtWidgets.QLineEdit(parent=self.groupBox)
        self.txttentv.setGeometry(QtCore.QRect(100, 80, 161, 22))

        # Tiêu đề
        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        self.label.setGeometry(QtCore.QRect(280, 10, 231, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet(
            "color: #4A90E2; background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #E1F5FE, stop:1 #B3E5FC);")

        # GroupBox xử lý
        self.groupBox_6 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_6.setGeometry(QtCore.QRect(620, 280, 181, 271))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setPointSize(10)
        self.groupBox_6.setFont(font)
        self.btnthem = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnthem.setGeometry(QtCore.QRect(40, 30, 93, 28))
        self.btnlammoi = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnlammoi.setGeometry(QtCore.QRect(40, 180, 93, 28))
        self.btnluu = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnluu.setGeometry(QtCore.QRect(40, 80, 93, 28))
        self.btnxoa = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnxoa.setGeometry(QtCore.QRect(40, 130, 93, 28))
        self.btnthoat = QtWidgets.QPushButton(parent=self.groupBox_6)
        self.btnthoat.setGeometry(QtCore.QRect(40, 230, 93, 28))

        # GroupBox danh sách
        self.groupBox_2 = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.groupBox_2.setGeometry(QtCore.QRect(10, 280, 601, 271))
        self.groupBox_2.setFont(font)
        self.tblds = QtWidgets.QTableWidget(parent=self.groupBox_2)
        self.tblds.setGeometry(QtCore.QRect(10, 20, 581, 241))
        self.tblds.setColumnCount(7)
        self.tblds.setHorizontalHeaderLabels([
            "Mã Thành Viên", "Họ Tên", "Ngày Sinh", "Giới Tính", "SĐT", "Địa Chỉ", "Mã CLB"
        ])

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Kết nối sự kiện
        self.btnthem.clicked.connect(self.add_account)
        self.btnxoa.clicked.connect(self.delete_account)
        self.btnluu.clicked.connect(self.update_account)
        self.btntimkiem.clicked.connect(self.search_account)
        self.btnxuat.clicked.connect(self.export_data)
        self.btnnhap.clicked.connect(self.import_data)
        self.btnlammoi.clicked.connect(self.lam_moi_du_lieu)
        self.btnthoat.clicked.connect(self.thoat)
        self.tblds.itemSelectionChanged.connect(self.table_item_selected)

        # Load dữ liệu ban đầu
        self.load_data()
        self.load_cbbclb()
        self.load_cbbmatv()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.groupBox_3.setTitle(_translate("MainWindow", "Tìm kiếm"))
        self.label_20.setText(_translate("MainWindow", "Tên thành viên:"))
        self.btnxuat.setText(_translate("MainWindow", "Xuất Excel"))
        self.btnnhap.setText(_translate("MainWindow", "Nhập Excel"))
        self.btntimkiem.setText(_translate("MainWindow", "Tìm kiếm"))
        self.groupBox.setTitle(_translate("MainWindow", "Thông tin thành viên:"))
        self.label_5.setText(_translate("MainWindow", "Mã thành viên"))
        self.label_11.setText(_translate("MainWindow", "Họ tên:"))
        self.label_15.setText(_translate("MainWindow", "Ngày sinh:"))
        self.label_17.setText(_translate("MainWindow", "Địa chỉ:"))
        self.label_18.setText(_translate("MainWindow", "Mã CLB:"))
        self.label_19.setText(_translate("MainWindow", "Số điện thoại: "))
        self.label_16.setText(_translate("MainWindow", "Giới tính:"))
        self.label.setText(_translate("MainWindow", "QUẢN LÝ THÀNH VIÊN"))
        self.groupBox_6.setTitle(_translate("MainWindow", "Xử lí"))
        self.btnthem.setText(_translate("MainWindow", "Thêm"))
        self.btnlammoi.setText(_translate("MainWindow", "Làm mới"))
        self.btnluu.setText(_translate("MainWindow", "Lưu"))
        self.btnxoa.setText(_translate("MainWindow", "Xóa"))
        self.btnthoat.setText(_translate("MainWindow", "Thoát"))
        self.groupBox_2.setTitle(_translate("MainWindow", "Danh sách thành viên:"))

    def load_data(self):
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM thanhvien")
            rows = cursor.fetchall()
            self.tblds.setRowCount(0)
            for row in rows:
                rowPosition = self.tblds.rowCount()
                self.tblds.insertRow(rowPosition)
                for col, value in enumerate(row):
                    self.tblds.setItem(rowPosition, col, QtWidgets.QTableWidgetItem(str(value)))
            cursor.close()
            connection.close()

    def load_cbbclb(self):
        try:
            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(self.MainWindow, "Lỗi", "Không thể kết nối đến cơ sở dữ liệu!")
                return
            cursor = connection.cursor()
            cursor.execute("SELECT ma_clb, ten_clb FROM CauLacBo")
            rows = cursor.fetchall()
            self.txtmaclb.clear()
            self.club_mapping = {}
            for row in rows:
                ma_clb, ten_clb = row
                display_text = f"{ma_clb} - {ten_clb}"
                self.club_mapping[display_text] = ma_clb
                self.txtmaclb.addItem(display_text)
            cursor.close()
            connection.close()
        except Exception as e:
            QMessageBox.critical(self.MainWindow, "Lỗi", f"Không thể tải danh sách CLB: {str(e)}")

    def load_cbbmatv(self):
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute("SELECT ho_ten FROM ThanhVien")
            rows = cursor.fetchall()
            self.txtmatv_2.clear()
            for row in rows:
                self.txtmatv_2.addItem(row[0])
            cursor.close()
            connection.close()

    def import_data(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(None, "Chọn file Excel", "", "Excel Files (*.xlsx)")
            if not file_path:
                return

            wb = load_workbook(file_path)
            ws = wb.active

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            inserted_count = 0
            duplicate_matvs = []
            invalid_clbs = []

            cursor.execute("SELECT ma_clb FROM CauLacBo")
            valid_clbs = {row[0] for row in cursor.fetchall()}

            for row in ws.iter_rows(min_row=2, values_only=True):
                matv, hoten, ngaysinh, gioitinh, sdt, diachi, maclb = row

                if not matv or not hoten:
                    continue

                cursor.execute("SELECT COUNT(*) FROM ThanhVien WHERE ma_tv = %s", (matv,))
                if cursor.fetchone()[0] > 0:
                    duplicate_matvs.append(str(matv))
                    continue

                if maclb and maclb not in valid_clbs:
                    invalid_clbs.append(str(maclb))
                    continue

                cursor.execute("""
                    INSERT INTO ThanhVien (ma_tv, ho_ten, ngay_sinh, gioi_tinh, sdt, dia_chi, ma_clb)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (matv, hoten, ngaysinh, gioitinh, sdt, diachi, maclb))
                inserted_count += 1

            connection.commit()

            message = f"Đã nhập thành công {inserted_count} thành viên từ Excel!"
            if duplicate_matvs:
                message += f"\nCác mã thành viên bị trùng (không được nhập): {', '.join(duplicate_matvs)}"
            if invalid_clbs:
                message += f"\nCác mã CLB không tồn tại (bản ghi bị bỏ qua): {', '.join(set(invalid_clbs))}"
            QMessageBox.information(None, "Kết quả nhập", message)

            self.load_data()
            self.load_cbbmatv()
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi nhập: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def add_account(self):
        try:
            matv = self.txtmatv.text().strip()
            hoten = self.txttentv.text().strip()
            ngaysinh = self.txtngaysinh.date().toString("yyyy-MM-dd")
            gioitinh = self.txtgioitinh.currentText().strip()
            sdt = self.txtsdt.text().strip()
            diachi = self.txtdiachi.text().strip()
            selected_text = self.txtmaclb.currentText().strip()
            maclb = self.club_mapping.get(selected_text, "")

            if not matv or not hoten:
                QMessageBox.warning(None, "Lỗi", "Mã thành viên và Họ tên không được để trống!")
                return

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM ThanhVien WHERE ma_tv = %s OR sdt = %s", (matv, sdt))
            count = cursor.fetchone()[0]

            if count > 0:
                QMessageBox.warning(None, "Lỗi", "Mã thành viên hoặc số điện thoại đã tồn tại!")
                cursor.close()
                connection.close()
                return

            cursor.execute("""
                INSERT INTO ThanhVien (ma_tv, ho_ten, ngay_sinh, gioi_tinh, sdt, dia_chi, ma_clb)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (matv, hoten, ngaysinh, gioitinh, sdt, diachi, maclb))
            connection.commit()
            QMessageBox.information(None, "Thành công", "Thêm thành viên thành công!")
            self.load_data()
            self.load_cbbmatv()
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi thêm: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def delete_account(self):
        try:
            row = self.tblds.currentRow()
            if row == -1:
                QMessageBox.warning(None, "Lỗi", "Vui lòng chọn một dòng để xóa!")
                return

            matv = self.tblds.item(row, 0).text()
            reply = QMessageBox.question(None, "Xác nhận", "Bạn có chắc muốn xóa thành viên này?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                connection = get_db_connection()
                if not connection:
                    QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                    return

                cursor = connection.cursor()
                cursor.execute("DELETE FROM ThanhVien WHERE ma_tv = %s", (matv,))
                connection.commit()
                QMessageBox.information(None, "Thành công", "Xóa thành viên thành công!")
                self.load_data()
                self.load_cbbmatv()
                cursor.close()
                connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi xóa: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def update_account(self):
        try:
            row = self.tblds.currentRow()
            if row == -1:
                QMessageBox.warning(None, "Lỗi", "Vui lòng chọn một dòng để sửa!")
                return

            matv = self.txtmatv.text().strip()
            hoten = self.txttentv.text().strip()
            ngaysinh = self.txtngaysinh.date().toString("yyyy-MM-dd")
            gioitinh = self.txtgioitinh.currentText().strip()
            sdt = self.txtsdt.text().strip()
            diachi = self.txtdiachi.text().strip()
            selected_text = self.txtmaclb.currentText().strip()
            maclb = self.club_mapping.get(selected_text, "")

            if not matv or not hoten:
                QMessageBox.warning(None, "Lỗi", "Mã thành viên và Họ tên không được để trống!")
                return

            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            cursor.execute("""
                UPDATE ThanhVien SET ho_ten = %s, ngay_sinh = %s, gioi_tinh = %s, sdt = %s, dia_chi = %s, ma_clb = %s
                WHERE ma_tv = %s
            """, (hoten, ngaysinh, gioitinh, sdt, diachi, maclb, matv))
            connection.commit()
            QMessageBox.information(None, "Thành công", "Cập nhật thành viên thành công!")
            self.load_data()
            self.load_cbbmatv()
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi sửa: {str(e)}")
            if 'connection' in locals():
                connection.rollback()
                cursor.close()
                connection.close()

    def search_account(self):
        try:
            hoten = self.txtmatv_2.currentText()
            connection = get_db_connection()
            if not connection:
                QMessageBox.critical(None, "Lỗi", "Không thể kết nối tới database!")
                return

            cursor = connection.cursor()
            cursor.execute("SELECT * FROM ThanhVien WHERE ho_ten = %s", (hoten,))
            rows = cursor.fetchall()
            self.tblds.setRowCount(0)
            for row in rows:
                rowPosition = self.tblds.rowCount()
                self.tblds.insertRow(rowPosition)
                for col, value in enumerate(row):
                    self.tblds.setItem(rowPosition, col, QtWidgets.QTableWidgetItem(str(value)))
            cursor.close()
            connection.close()

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi tìm kiếm: {str(e)}")

    def export_data(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ThanhVien Data"
            headers = ["Mã Thành Viên", "Họ Tên", "Ngày Sinh", "Giới Tính", "SĐT", "Địa Chỉ", "Mã CLB"]
            ws.append(headers)

            for row in range(self.tblds.rowCount()):
                row_data = []
                for col in range(self.tblds.columnCount()):
                    item = self.tblds.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            file_path, _ = QFileDialog.getSaveFileName(None, "Lưu file Excel", "", "Excel Files (*.xlsx)")
            if file_path:
                wb.save(file_path)
                QMessageBox.information(None, "Thành công", "Dữ liệu đã được xuất thành công!")

        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi xuất Excel: {str(e)}")

    def lam_moi_du_lieu(self):
        try:
            self.txtmatv.clear()
            self.txttentv.clear()
            self.txtgioitinh.setCurrentIndex(0)
            self.txtngaysinh.setDate(QtCore.QDate.currentDate())
            self.txtsdt.clear()
            self.txtdiachi.clear()
            self.txtmaclb.setCurrentIndex(-1)
            self.txtmatv_2.setCurrentIndex(-1)
            self.load_data()
            QMessageBox.information(None, "Thành công", "Dữ liệu đã được làm mới!")
        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Lỗi khi làm mới: {str(e)}")

    def thoat(self):
        reply = QMessageBox.question(None, "Xác nhận", "Bạn có muốn thoát không?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.MainWindow.close()

    def table_item_selected(self):
        selected_row = self.tblds.currentRow()
        if selected_row != -1:
            self.txtmatv.setText(self.tblds.item(selected_row, 0).text())
            self.txttentv.setText(self.tblds.item(selected_row, 1).text())
            date_str = self.tblds.item(selected_row, 2).text()
            date = QtCore.QDate.fromString(date_str, "yyyy-MM-dd")
            self.txtngaysinh.setDate(date)
            gender = self.tblds.item(selected_row, 3).text()
            self.txtgioitinh.setCurrentText(gender)
            self.txtsdt.setText(self.tblds.item(selected_row, 4).text())
            self.txtdiachi.setText(self.tblds.item(selected_row, 5).text())
            ma_clb = self.tblds.item(selected_row, 6).text()
            for i in range(self.txtmaclb.count()):
                item_text = self.txtmaclb.itemText(i)
                if item_text.startswith(ma_clb):
                    self.txtmaclb.setCurrentIndex(i)
                    break


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())